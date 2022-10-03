import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
file_name = "C:\\Users\\Venkat\\Desktop\\testExcelFiles\\Viper PW worksheet.xlsx"
sheet_name = "Main Creditors"

def parser_merged_cell(sheet: Worksheet, row, col):
    """
    Check whether it is a merged cell and get the value of the corresponding row and column cell.
    If it is a merged cell, take the value of the cell in the upper left corner of the merged area as the value of the current cell; otherwise, directly return the value of the cell
    : param sheet: current sheet object
    : param row: the row of the cell to be obtained
    : param col: the column of the cell to be obtained
    :return: 
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance (cell, MergedCell): # judge whether the cell is a merged cell
        for merged_range in sheet.merged_cell_Ranges: # loop to find the merge range to which the cell belongs
            if cell.coordinate in merged_range:
                #Gets the cell in the upper left corner of the merge range and returns it as the value of the cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


if __name__ == "__main__":
    wb = xl.load_workbook(file_name)
    sheet_ = wb[sheet_name]
    max_columns_count = len(set(sheet_.columns))
    max_rows_count = len(set(sheet_.rows))
    for row_index in range(1, max_rows_count):
        for col_index in range(1, max_columns_count):
            cell_ = parser_merged_cell(sheet_, row_index, col_index)
            print("line {0}, column% {1}: {2}".format(row_index, col_index, cell_.value))